import requests
import json
import jsonpath
import openpyxl


class Common:

    def __init__(self, FileNamePath, SheetName):
        global wb
        global sh
        wb = openpyxl.load_workbook(FileNamePath)
        sh = wb[SheetName]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        column = sh.max_column
        return column

    def fetch_key_name(self):
        c = sh.max_column
        li = []
        for i in range(1, c + 1):
            cell = sh.cell(row=1, column=i)
            li.insert(i - 1, cell.value)

        return li

    def update_request_with_date(self, rowNumber, jsonRequest, keyList):
        c = sh.max_column  # column 4
        for i in range(1, c + 1):
            cell = sh.cell(row=rowNumber, column=i)
            jsonRequest[keyList[-1]] = cell.value

        return jsonRequest
